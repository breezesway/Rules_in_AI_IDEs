"""
Read Q10 (multi-select) from English/Chinese survey Excel files under survey/survey_results,
map Chinese options to English labels, merge into one CSV; Other free-text in a separate column.

Requires: pip install openpyxl
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

try:
    import openpyxl
except ImportError as e:  # pragma: no cover
    raise SystemExit("Install openpyxl first: pip install openpyxl") from e

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent.parent
SURVEY_DIR = REPO_ROOT / "survey" / "survey_results"
FILE_EN = SURVEY_DIR / "Survey on AI IDE Rules_English.xlsx"
FILE_ZH = SURVEY_DIR / "Survey on AI IDE Rules_Chinese.xlsx"
OUTPUT_CSV = SCRIPT_DIR / "q10_combined_responses.csv"

# Column names aligned with tier_1 / English survey options (fixed order for downstream stats)
OPTION_LABELS: Tuple[str, ...] = (
    "Correction",
    "Refinement",
    "Synchronization",
    "Context Enrichment",
    "Expansion",
    "Pruning",
)

# English Q10 cell option prefixes (match longer phrases first)
_EN_PREFIXES = (
    "CONTEXT ENRICHMENT",
    "SYNCHRONIZATION",
    "REFINEMENT",
    "CORRECTION",
    "EXPANSION",
    "PRUNING",
    "OTHER",
)
_EN_PATTERN = re.compile(
    r"\s*("
    + "|".join(re.escape(p) for p in sorted(_EN_PREFIXES, key=len, reverse=True))
    + r")\s*:",
    re.IGNORECASE,
)
_EN_PREFIX_TO_LABEL = {
    "CORRECTION": "Correction",
    "REFINEMENT": "Refinement",
    "SYNCHRONIZATION": "Synchronization",
    "CONTEXT ENRICHMENT": "Context Enrichment",
    "EXPANSION": "Expansion",
    "PRUNING": "Pruning",
    "OTHER": "Other",
}


def _norm_en_prefix(raw: str) -> str:
    return " ".join(raw.upper().split())


def parse_english_q10(cell: object) -> Tuple[Dict[str, bool], str]:
    """Parse English Q10 single-cell text; return {English label: selected} and Other free text."""
    if cell is None:
        return {}, ""
    text = str(cell).strip()
    if not text:
        return {}, ""

    selected: Dict[str, bool] = {}
    other_text = ""
    matches = list(_EN_PATTERN.finditer(text))
    for i, m in enumerate(matches):
        prefix = _norm_en_prefix(m.group(1))
        label = _EN_PREFIX_TO_LABEL.get(prefix)
        if label is None:
            continue
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        body = text[start:end].strip().rstrip(",").strip()
        selected[label] = True
        if label == "Other":
            other_text = body
    return selected, other_text


@dataclass
class ChineseQ10Columns:
    option_cols: Dict[str, int]  # label -> 0-based column index
    other_flag_col: Optional[int]
    other_text_col: Optional[int]


def detect_chinese_q10_columns(header: Tuple[Optional[object], ...]) -> ChineseQ10Columns:
    """Detect Q10 option columns and Other/fill-in columns from Chinese headers."""
    option_cols: Dict[str, int] = {}
    other_flag: Optional[int] = None
    other_text: Optional[int] = None

    for j, raw in enumerate(header):
        if raw is None:
            continue
        h = str(raw)
        if "Q10" not in h:
            continue
        if "纠错" in h:
            option_cols["Correction"] = j
        elif "优化" in h and "规则" in h:
            option_cols["Refinement"] = j
        elif "同步" in h and "代码结构" in h:
            option_cols["Synchronization"] = j
        elif "补充上下文" in h:
            option_cols["Context Enrichment"] = j
        elif "扩展" in h and "业务模块" in h:
            option_cols["Expansion"] = j
        elif "清理" in h and "过时" in h:
            option_cols["Pruning"] = j
        elif "其他" in h and "选项填空" in h:
            other_text = j
        elif "其他" in h:
            other_flag = j

    missing = [lab for lab in OPTION_LABELS if lab not in option_cols]
    if missing:
        raise ValueError(f"Could not detect Q10 option columns from Chinese headers: {missing}")
    return ChineseQ10Columns(
        option_cols=option_cols,
        other_flag_col=other_flag,
        other_text_col=other_text,
    )


def _cell_truthy(v: object) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and not v.strip():
        return False
    return True


def parse_chinese_q10_row(
    row: Tuple[Optional[object], ...], spec: ChineseQ10Columns
) -> Tuple[Dict[str, bool], str]:
    selected: Dict[str, bool] = {}
    for lab in OPTION_LABELS:
        j = spec.option_cols[lab]
        if j < len(row) and _cell_truthy(row[j]):
            selected[lab] = True

    other_text = ""
    if spec.other_text_col is not None and spec.other_text_col < len(row):
        v = row[spec.other_text_col]
        if isinstance(v, str) and v.strip():
            other_text = v.strip()

    other_flag = False
    if spec.other_flag_col is not None and spec.other_flag_col < len(row):
        other_flag = _cell_truthy(row[spec.other_flag_col])
    if other_text:
        other_flag = True

    if other_flag:
        selected["Other"] = True
    return selected, other_text


def _find_english_q10_col(header: Tuple[Optional[object], ...]) -> int:
    for j, raw in enumerate(header):
        if raw is None:
            continue
        s = str(raw)
        if "Q10" in s and "modify" in s.lower():
            return j
    raise ValueError("Q10 (modify existing rule) column not found in English headers")


def _load_sheet(path: Path, sheet_hint_zh: bool):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_hint_zh:
        if "同步数据表" in wb.sheetnames:
            ws = wb["同步数据表"]
        else:
            ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    data_rows: List[Tuple] = list(rows_iter)
    wb.close()
    return header, data_rows


def detect_timestamp_col(header: Tuple[Optional[object], ...]) -> int:
    """English often uses 时间戳记; Tencent Chinese exports use 开始答题时间; avoid treating 编号 as timestamp."""
    for j, raw in enumerate(header):
        if raw is None:
            continue
        s = str(raw).strip()
        if s == "时间戳记" or "开始答题时间" in s or "提交答卷时间" in s:
            return j
    return 0


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def iter_combined_records() -> Iterable[Dict[str, object]]:
    header_en, rows_en = _load_sheet(FILE_EN, sheet_hint_zh=False)
    q10_en = _find_english_q10_col(header_en)
    ts_en = detect_timestamp_col(header_en)

    header_zh, rows_zh = _load_sheet(FILE_ZH, sheet_hint_zh=True)
    zh_spec = detect_chinese_q10_columns(header_zh)
    ts_zh = detect_timestamp_col(header_zh)

    for i, row in enumerate(rows_en, start=2):
        sel, other_txt = parse_english_q10(row[q10_en] if q10_en < len(row) else None)
        ts = _cell_str(row[ts_en]) if ts_en < len(row) else ""
        rec: Dict[str, object] = {
            "source": "english",
            "sheet_row": i,
            "timestamp": ts,
        }
        for lab in OPTION_LABELS:
            rec[lab] = 1 if sel.get(lab) else 0
        rec["Other_selected"] = 1 if sel.get("Other") else 0
        rec["other_text"] = other_txt
        yield rec

    for i, row in enumerate(rows_zh, start=2):
        sel, other_txt = parse_chinese_q10_row(row, zh_spec)
        ts = _cell_str(row[ts_zh]) if ts_zh < len(row) else ""
        rec = {
            "source": "chinese",
            "sheet_row": i,
            "timestamp": ts,
        }
        for lab in OPTION_LABELS:
            rec[lab] = 1 if sel.get(lab) else 0
        rec["Other_selected"] = 1 if sel.get("Other") else 0
        rec["other_text"] = other_txt
        yield rec


def _print_summary(records: List[Dict[str, object]]) -> None:
    n = len(records)
    print(f"Combined responses: {n} (english + chinese)")
    for lab in OPTION_LABELS:
        c = sum(int(r[lab]) for r in records)
        print(f"  {lab}: {c} ({100.0 * c / n:.1f}% of respondents)" if n else f"  {lab}: {c}")
    oc = sum(int(r["Other_selected"]) for r in records)
    print(f"  Other (any): {oc}")
    with_other_text = sum(1 for r in records if str(r.get("other_text", "")).strip())
    print(f"  other_text non-empty: {with_other_text}")


def main() -> None:
    if not FILE_EN.is_file():
        raise FileNotFoundError(FILE_EN)
    if not FILE_ZH.is_file():
        raise FileNotFoundError(FILE_ZH)

    records = list(iter_combined_records())
    fieldnames = [
        "source",
        "sheet_row",
        "timestamp",
        *OPTION_LABELS,
        "Other_selected",
        "other_text",
    ]
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for rec in records:
            w.writerow(rec)

    print(f"Wrote: {OUTPUT_CSV}")
    _print_summary(records)


if __name__ == "__main__":
    main()
